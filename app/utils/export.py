import io
import csv
import json
from collections import defaultdict
from datetime import datetime


def _get_field_value(record, field_id):
    rv = record.values.filter_by(entity_field_id=field_id).first()
    return rv.value_text if rv else ''


def _get_numeric_value(record, field_id):
    rv = record.values.filter_by(entity_field_id=field_id).first()
    return rv.value_number if rv else None


def _base_headers(fields):
    return (
        ['record_id', 'display_label', 'org_unit', 'org_unit_path',
         'parent_record_id', 'created_by', 'created_at', 'updated_at']
        + [f.label for f in fields]
    )


def _base_row(record, fields):
    return (
        [
            record.id,
            record.display_label or '',
            record.org_unit.name if record.org_unit else '',
            record.org_unit.path if record.org_unit else '',
            record.parent_record_id or '',
            record.creator.username if record.creator else '',
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '',
            record.updated_at.strftime('%Y-%m-%d %H:%M') if record.updated_at else '',
        ]
        + [_get_field_value(record, f.id) for f in fields]
    )


# ── Record-level CSV ──────────────────────────────────────────────────────────

def export_records_csv(records, entity_type, fields):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_base_headers(fields))
    for rec in records:
        writer.writerow(_base_row(rec, fields))
    output.seek(0)
    return output


# ── Record-level Excel ────────────────────────────────────────────────────────

def export_records_excel(records, entity_type, fields):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl is required for Excel export')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity_type.name[:31]

    headers = _base_headers(fields)
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1E40AF')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left')

    for rec in records:
        ws.append(_base_row(rec, fields))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Aggregate CSV ─────────────────────────────────────────────────────────────

def export_aggregate_csv(records, entity_type):
    from app.models.entity import EntityField
    numeric_fields = EntityField.query.filter_by(
        entity_type_id=entity_type.id, field_type='number'
    ).all()

    # Group by org_unit + YYYY-MM
    buckets = defaultdict(lambda: {'count': 0, 'sums': defaultdict(float)})
    for rec in records:
        period = rec.created_at.strftime('%Y-%m') if rec.created_at else 'unknown'
        unit_name = rec.org_unit.name if rec.org_unit else ''
        unit_path = rec.org_unit.path if rec.org_unit else ''
        key = (period, unit_name, unit_path)
        buckets[key]['count'] += 1
        for nf in numeric_fields:
            val = _get_numeric_value(rec, nf.id)
            if val is not None:
                buckets[key]['sums'][nf.name] += val

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        ['period', 'org_unit', 'org_unit_path', 'record_count']
        + [f.label for f in numeric_fields]
    )
    for (period, unit_name, unit_path), data in sorted(buckets.items()):
        writer.writerow(
            [period, unit_name, unit_path, data['count']]
            + [data['sums'].get(nf.name, '') for nf in numeric_fields]
        )
    output.seek(0)
    return output


# ── Aggregate Excel ───────────────────────────────────────────────────────────

def export_aggregate_excel(records, entity_type):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl is required for Excel export')

    from app.models.entity import EntityField
    numeric_fields = EntityField.query.filter_by(
        entity_type_id=entity_type.id, field_type='number'
    ).all()

    buckets = defaultdict(lambda: {'count': 0, 'sums': defaultdict(float)})
    for rec in records:
        period = rec.created_at.strftime('%Y-%m') if rec.created_at else 'unknown'
        unit_name = rec.org_unit.name if rec.org_unit else ''
        unit_path = rec.org_unit.path if rec.org_unit else ''
        key = (period, unit_name, unit_path)
        buckets[key]['count'] += 1
        for nf in numeric_fields:
            val = _get_numeric_value(rec, nf.id)
            if val is not None:
                buckets[key]['sums'][nf.name] += val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Aggregate'
    headers = ['Period', 'Org Unit', 'Path', 'Record Count'] + [f.label for f in numeric_fields]
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='166534')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)

    for (period, unit_name, unit_path), data in sorted(buckets.items()):
        ws.append(
            [period, unit_name, unit_path, data['count']]
            + [data['sums'].get(nf.name, '') for nf in numeric_fields]
        )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── DHIS2 JSON ────────────────────────────────────────────────────────────────

def export_dhis2_json(records, entity_type):
    """
    Produces a DHIS2-compatible dataValueSets JSON.
    One dataValue per entity_type per org_unit per month = record count.
    Numeric fields produce additional dataValues with their sums.
    """
    from app.models.entity import EntityField
    numeric_fields = EntityField.query.filter_by(
        entity_type_id=entity_type.id, field_type='number'
    ).all()

    buckets = defaultdict(lambda: {'count': 0, 'sums': defaultdict(float)})
    for rec in records:
        period_str = rec.created_at.strftime('%Y%m') if rec.created_at else '000000'
        unit_ref = (rec.org_unit.code or rec.org_unit.name) if rec.org_unit else 'UNKNOWN'
        key = (period_str, unit_ref)
        buckets[key]['count'] += 1
        for nf in numeric_fields:
            val = _get_numeric_value(rec, nf.id)
            if val is not None:
                buckets[key]['sums'][nf.name] += val

    data_values = []
    for (period, unit_ref), data in sorted(buckets.items()):
        data_values.append({
            'dataElement': f'{entity_type.slug}_count',
            'period': period,
            'orgUnit': unit_ref,
            'value': str(data['count']),
            'comment': f'{entity_type.name} record count',
        })
        for nf in numeric_fields:
            if nf.name in data['sums']:
                data_values.append({
                    'dataElement': f'{entity_type.slug}_{nf.name}',
                    'period': period,
                    'orgUnit': unit_ref,
                    'value': str(round(data['sums'][nf.name], 4)),
                    'comment': f'{entity_type.name} — {nf.label} sum',
                })

    payload = {
        'dataValueSets': data_values,
        'meta': {
            'exported_at': datetime.utcnow().isoformat(),
            'entity_type': entity_type.name,
            'record_count': len(records),
        }
    }
    return json.dumps(payload, indent=2)
